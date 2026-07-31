import re
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_json_from_output(text):
    """Extrahiert ein JSON-Objekt aus einem Text."""
    try:
        # Escaped backticks to prevent markdown parser breaks
        match = re.search(r'\`\`\`(?:json)?\s*(\{.*?\})\s*\`\`\`', text, re.DOTALL)
        if match:
            return json.loads(match.group(1))
        # Fallback für Listen (falls das JSON ein Array ist)
        match = re.search(r'\`\`\`(?:json)?\s*(\[.*?\])\s*\`\`\`', text, re.DOTALL)
        if match:
            return json.loads(match.group(1))
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON Parse-Fehler. Rohdaten:\n{text}")
        return None

def extract_known_canon(full_json_data):
    """Erstellt das Lexikon aller bekannten SAP Merkmalsnamen."""
    canon = set()
    classes = full_json_data.get("classes", [])
    for cls in classes:
        for char in cls.get("characteristics", []):
            char_name = char.get("name")
            if char_name:
                canon.add(char_name.upper())
    return canon

def optimize_sap_logic(syntax, known_canon):
    """Kürzt die Logik via Regex und wendet Ghost-Logic an (mit Schutzschilden für Werte)."""
    bad_patterns = [r"specified\s+dummy", r"dummy\s*=\s*'yes'"]
    has_ghost = any(re.search(p, syntax, re.IGNORECASE) for p in bad_patterns)
    
    clean_syntax = syntax
    for p in bad_patterns:
        clean_syntax = re.sub(p, "True", clean_syntax, flags=re.IGNORECASE)

    # ==========================================
    # NEU: SCHUTZSCHILDE FÜR STRINGS & SYS-VARS
    # ==========================================
    placeholders = {}

    # 1. Strings schützen ('value')
    def repl_string(m):
        key = f"__STR_{len(placeholders)}__"
        placeholders[key] = m.group(0)
        return key
    clean_syntax = re.sub(r"'[^']*'", repl_string, clean_syntax)

    # 2. Systemvariablen schützen ($self, .me, etc.)
    def repl_sys(m):
        key = f"__SYS_{len(placeholders)}__"
        placeholders[key] = m.group(0)
        return key
    clean_syntax = re.sub(r'(\$|\.)[A-Za-z_][A-Za-z0-9_]*', repl_sys, clean_syntax)
        
    tokens = re.split(r'(\W+)', clean_syntax)
    optimized_tokens = []
    
    # Erweiterte Ausnahmeliste (jetzt inkl. SAP Operatoren)
    allowed_keywords = {
        "AND", "OR", "NOT", "IF", "IN", "TRUE", "FALSE",
        "NE", "EQ", "LT", "GT", "LE", "GE", "SPECIFIED"
    }
    
    for t in tokens:
        if re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', t):
            # Wenn es ein Schutzschild ist, ignorieren wir es
            if t.startswith("__STR_") or t.startswith("__SYS_"):
                optimized_tokens.append(t)
            elif t.upper() not in known_canon and t.upper() not in allowed_keywords:
                optimized_tokens.append(f"[UNCLASSIFIED_{t}]")
            else:
                optimized_tokens.append(t)
        else:
            optimized_tokens.append(t)
            
    final_syntax = "".join(optimized_tokens).strip()
    
    # ==========================================
    # SCHUTZSCHILDE WIEDER AUFLÖSEN
    # ==========================================
    for key, val in placeholders.items():
        final_syntax = final_syntax.replace(key, val)
    
    if has_ghost:
        final_syntax += " /* GHOST-LOGIC: Enthält unklassifizierte oder Dummy-Bedingungen */"
        
    return final_syntax

def heavy_preprocessing(char_dict, known_canon):
    """Die Waschmaschine: Entfernt Kommentare und reinigt Strings vor der Analyse."""
    char_copy = dict(char_dict)
    values = char_copy.get("values", [])
    
    optimized_values = []
    for val in values:
        deps = val.get("dependency", [])
        if not deps:
            optimized_values.append(val)
            continue
            
        syntaxes = []
        for d in deps:
            raw_syntax = d.get("syntax", "")
            no_comments = re.sub(r'\*.*?(?=\n|$)', '', raw_syntax)
            no_comments = re.sub(r'/\*.*?\*/', '', no_comments, flags=re.DOTALL)
            clean_str = " ".join(no_comments.split())
            if clean_str:
                syntaxes.append(clean_str)
                
        if not syntaxes:
            optimized_values.append(val)
            continue
            
        combined_syntax = " AND ".join(f"({s})" for s in syntaxes)
        optimized_syntax = optimize_sap_logic(combined_syntax, known_canon)
        
        val["dependency"] = [{"precondition": "OPTIMIZED_AST", "syntax": optimized_syntax}]
        optimized_values.append(val)
        
    char_copy["values"] = optimized_values
    return char_copy

# ==========================================
# SUB-CLUSTERING TABLE BUILDER (Option A + Sub-Signatures + Tolerant Regex)
# ==========================================
def create_heuristic_table_plan(char_dict):
    char_name = char_dict.get("name", "Unknown")
    values = char_dict.get("values", [])

    groups = {} 
    standalone_items = []
    bad_operators = ["<", ">", " ne ", " or ", " | ", "never_possible"]

    for val in values:
        code = val.get("code")
        deps = val.get("dependency", [])
        if not deps:
            continue

        logic = deps[0].get("syntax", "")
        original_item = {"wert": code, "logik": logic}

        # 1. Sofortiger Ausschluss bei komplexen Operatoren
        logic_lower = logic.lower()
        if any(op in logic_lower for op in bad_operators):
            standalone_items.append(original_item)
            continue

        # 2. Parsen der Spalten
        # Zuerst nur den Ghost-Logic-Kommentar wegwerfen
        clean_logic = re.sub(r'/\*.*?\*/', '', logic).strip()
        
        # Aufsplitten am AND-Operator (Dies erzeugt die Dangling Brackets)
        parts = [p.strip() for p in re.split(r'\s+AND\s+|\s+&\s+', clean_logic, flags=re.IGNORECASE)]

        row_dict = {}
        is_valid_row = True
        
        for part in parts:
            # Hilfsstring für die reine TRUE/FALSE/SPECIFIED Prüfung (Klammern entfernt)
            clean_part = part.strip('() \t').upper()
            if not clean_part or clean_part in ["ALWAYS_TRUE", "TRUE", "FALSE"]:
                continue
                
            # Regex 1: Prüfe auf Variable = 'Wert' (Tolerant gegenüber Dangling Brackets: ^\(* und \)*$)
            eq_match = re.match(r'^\(*\s*([A-Za-z0-9_]+)\s*=\s*(.*?)\s*\)*$', part)
            if eq_match:
                var = eq_match.group(1).upper()
                row_dict[var] = eq_match.group(2).replace("'", "").strip()
                continue
                
            # Regex 2: Prüfe auf Variable IN ('Wert1', 'Wert2') (Tolerant gegenüber Dangling Brackets)
            # Das innere (.+?) fängt sauber die '3','4' ab, während das äußere \)* die Restklammern schluckt.
            in_match = re.match(r'^\(*\s*([A-Za-z0-9_]+)\s+IN\s*\((.+?)\)\s*\)*$', part, flags=re.IGNORECASE)
            if in_match:
                var = in_match.group(1).upper()
                row_dict[var] = in_match.group(2).replace("'", "").replace(" ", "")
                continue
                
            # GHOST-LOGIC Überbleibsel tolerieren
            if "SPECIFIED" in clean_part:
                continue

            # Wenn es weder '=' noch 'IN' noch 'SPECIFIED' war, ist es zu komplex
            is_valid_row = False
            break

        # 3. Zuweisung in Signatur-Buckets
        if is_valid_row and row_dict:
            signature = tuple(sorted(row_dict.keys()))
            if signature not in groups:
                groups[signature] = []
            groups[signature].append({"code": code, "dict": row_dict, "logik": logic})
        else:
            standalone_items.append(original_item)

    # 4. JSON Outputs generieren
    plans = []
    table_counter = 1

    for sig, items in groups.items():
        if len(items) >= 2:
            cols = list(sig)
            plan = {
                "zielmerkmal": f"{char_name}_{table_counter}" if len(groups) > 1 or standalone_items else char_name,
                "decision": "TABLE",
                "bedingungs_spalten": cols,
                "zeilen_daten": [],
                "if_constraints_text": "Sub-Clustering: Homogene Spalten-Signatur erkannt.",
                "original_code_mapping": []
            }
            for item in items:
                row = [item["dict"].get(c, "") for c in cols] + [item["code"]]
                plan["zeilen_daten"].append(row)
                plan["original_code_mapping"].append({"wert": item["code"], "logik": item["logik"]})
            plans.append(plan)
            table_counter += 1
        else:
            standalone_items.append({"wert": items[0]["code"], "logik": items[0]["logik"]})

    if standalone_items:
        plans.append({
            "zielmerkmal": f"{char_name}_STANDALONE" if table_counter > 1 else char_name,
            "decision": "NO_TABLE",
            "bedingungs_spalten": [],
            "zeilen_daten": [],
            "if_constraints_text": "Sub-Clustering: Asymmetrische oder komplexe Ausreißer -> Standalone.",
            "original_code_mapping": standalone_items
        })

    return plans

def create_excel_dashboard(output_file, results_dict):
    """Erstellt das professionelle Entwickler-Dashboard in Excel."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    subheader_font = Font(name="Arial", size=12, bold=True, color="003366")
    text_font = Font(name="Arial", size=10)
    code_font = Font(name="Courier New", size=10, color="000000")
    code_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for char_name, data in results_dict.items():
        ws = wb.create_sheet(title=char_name[:31]) 
        
        ws.merge_cells('A1:G1')
        ws['A1'] = f"MAS MIGRATIONS-VORSCHLAG: {char_name}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 25

        ws['A3'] = "1. Herleitung (Gedankengang des Architekten)"
        ws['A3'].font = subheader_font
        ws['A4'] = data.get("herleitung", "Keine Herleitung verfügbar.")
        ws['A4'].font = text_font
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A4:G6')

        ws['A8'] = "2. Visuelle Variantentabelle (Für CU60)"
        ws['A8'].font = subheader_font
        
        start_row = 9
        headers = data.get("tabelle_kopf", [])
        for c_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=c_idx+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = table_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
        rows = data.get("tabelle_zeilen", [])
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + 1 + r_idx, column=c_idx+1, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

        code_start_row = start_row + len(rows) + 3
        ws.cell(row=code_start_row, column=1, value="3. Generierter AVC Constraint Code").font = subheader_font
        
        code_cell = ws.cell(row=code_start_row+1, column=1, value=data.get("avc_code", "// Kein Code generiert."))
        code_cell.font = code_font
        code_cell.fill = code_fill
        code_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=code_start_row+1, start_column=1, end_row=code_start_row+15, end_column=7)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20

    wb.save(output_file)
    print(f"📊 Excel Dashboard erfolgreich gespeichert unter: {output_file}")