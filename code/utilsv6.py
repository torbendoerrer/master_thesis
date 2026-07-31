import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def extract_json_from_output(text):
    """Extrahiert das JSON-Objekt sicher aus dem LLM Output, auch wenn Markdown drumherum ist."""
    try:
        # Dynamische Konstruktion des Suchmusters, um Parser-Probleme zu vermeiden
        pattern = r'`' * 3 + r'(?:json)?\s*(\{.*?\})\s*' + r'`' * 3
        match = re.search(pattern, text, re.DOTALL)
        if match:
            return json.loads(match.group(1))
        # Fallback: Versuche den ganzen Text zu parsen
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON Parse-Fehler. Rohdaten:\n{text}")
        return None

def create_excel_dashboard(output_file, results_dict):
    """Erstellt das professionelle Entwickler-Dashboard in Excel."""
    wb = openpyxl.Workbook()
    # Standard-Sheet löschen
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Styling definieren
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    subheader_font = Font(name="Arial", size=12, bold=True, color="003366")
    text_font = Font(name="Arial", size=10)
    code_font = Font(name="Courier New", size=10, color="000000")
    code_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for char_name, data in results_dict.items():
        # Excel Tab-Namen dürfen max 31 Zeichen lang sein
        ws = wb.create_sheet(title=char_name[:31]) 
        
        # 1. Titel
        ws.merge_cells('A1:G1')
        ws['A1'] = f"MAS MIGRATIONS-VORSCHLAG: {char_name}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 25

        # 2. Herleitung
        ws['A3'] = "1. Herleitung (Gedankengang des Architekten)"
        ws['A3'].font = subheader_font
        ws['A4'] = data.get("herleitung", "Keine Herleitung verfügbar.")
        ws['A4'].font = text_font
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A4:G6')

        # 3. Tabelle
        ws['A8'] = "2. Visuelle Variantentabelle (Für CU60)"
        ws['A8'].font = subheader_font
        
        start_row = 9
        headers = data.get("tabelle_kopf", [])
        for c_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=c_idx+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = table_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
        rows = data.get("tabelle_zeilen", [])
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + 1 + r_idx, column=c_idx+1, value=str(value))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

        # 4. Code
        code_start_row = start_row + len(rows) + 3
        ws.cell(row=code_start_row, column=1, value="3. Generierter AVC Constraint Code").font = subheader_font
        
        code_cell = ws.cell(row=code_start_row+1, column=1, value=data.get("avc_code", "// Kein Code generiert."))
        code_cell.font = code_font
        code_cell.fill = code_fill
        code_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=code_start_row+1, start_column=1, end_row=code_start_row+15, end_column=7)

        # Spaltenbreiten anpassen
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20

    wb.save(output_file)
    print(f"📊 Excel Dashboard erfolgreich gespeichert unter: {output_file}")