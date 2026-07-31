import json
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sympy
from sympy.logic.boolalg import simplify_logic, true, false

def extract_json_from_output(text):
    """Extrahiert das JSON-Objekt sicher aus dem LLM Output."""
    try:
        pattern = r'`' * 3 + r'(?:json)?\s*(\{.*?\})\s*' + r'`' * 3
        match = re.search(pattern, text, re.DOTALL)
        if match:
            return json.loads(match.group(1))
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"⚠️ JSON Parse-Fehler. Rohdaten:\n{text}")
        return None

# ==========================================
# 1. KNOWN CANON (Vokabel-Sammler)
# ==========================================
def extract_known_canon(data):
    """Extrahiert alle klassifizierten Merkmale automatisch aus dem JSON-Modell."""
    known_canon = set()
    for cls in data.get("classes", []):
        for char in cls.get("characteristics", []):
            name = char.get("name")
            if name:
                known_canon.add(name.upper())
    return known_canon

# ==========================================
# 2. BOOLEAN AST OPTIMIZER (Der "Compiler")
# ==========================================
def optimize_sap_logic(syntax_string, known_canon):
    """
    Säubert SAP Syntax, maskiert, wendet De Morgan an, löst Unclassified-Logik auf 
    und dokumentiert die Auflösung transparent im String.
    """
    if not syntax_string or syntax_string.strip() == "":
        return ""

    # --- 1. DIE WASCHMASCHINE (Sanitization) ---
    # Kommentare (*) zeilenweise entfernen und \n durch Leerzeichen ersetzen
    lines = syntax_string.split('\n')
    cleaned_lines = []
    for line in lines:
        if '*' in line:
            line = line.split('*')[0]  # Schneidet alles ab dem '*' ab
        cleaned_lines.append(line.strip())
    
    cleaned_string = " ".join(cleaned_lines)
    cleaned_string = re.sub(r'\s+', ' ', cleaned_string).strip()

    if not cleaned_string:
        return ""

    # --- 2. Tobi's Regeln für Geister-Merkmale taggen ---
    def tag_unclassified(match):
        word = match.group(0)
        keywords = {"and", "or", "not", "if", "in", "specified", "ne", "eq", "gt", "lt", "ge", "le", "dummy"}
        if word.lower() in keywords or word.isdigit():
            return word
        if word.upper() not in known_canon:
            return f"[UNCLASSIFIED_{word}]"
        return word
    
    parts = cleaned_string.split("'")
    for i in range(0, len(parts), 2):
        parts[i] = re.sub(r'\b[A-Za-z_][A-Za-z0-9_]*\b', tag_unclassified, parts[i])
    tagged_syntax = "'".join(parts)

    mask_map = {}
    reverse_map = {}
    ghost_notes = []
    counter = 0

    # --- 3. Maskierung für SymPy (Konditionen -> C0, C1) ---
    condition_pattern = r'((?:not\s+)?specified\s+(?:\[UNCLASSIFIED_\w+\]|\b[A-Za-z_][A-Za-z0-9_]*\b)|(?:\[UNCLASSIFIED_\w+\]|\b[A-Za-z_][A-Za-z0-9_]*\b)\s*(?:=|<>|!=|ne|eq|<|>|>=|<=|in)\s*(?:\'.*?\'|\(.*?\)))'
    
    def mask_condition(match):
        nonlocal counter
        cond = match.group(0)
        var_name = f"C{counter}"
        mask_map[var_name] = cond
        reverse_map[sympy.Symbol(var_name)] = cond
        counter += 1
        return var_name

    masked_syntax = re.sub(condition_pattern, mask_condition, tagged_syntax, flags=re.IGNORECASE)
    
    # --- 4. Geister-Merkmale vor SymPy zu True/False auswerten ---
    for var_name, original_cond in mask_map.items():
        if "[UNCLASSIFIED" in original_cond:
            if re.match(r'^not\s+specified', original_cond, flags=re.IGNORECASE):
                bool_val = "True"
                ghost_notes.append(f"{original_cond} -> TRUE")
            elif re.match(r'^specified', original_cond, flags=re.IGNORECASE):
                bool_val = "False"
                ghost_notes.append(f"{original_cond} -> FALSE")
            else:
                bool_val = "True"
                ghost_notes.append(f"{original_cond} -> TRUE")
                
            # Ersetze die Variable (z.B. C0) direkt durch True/False
            masked_syntax = re.sub(fr'\b{var_name}\b', bool_val, masked_syntax)
            if sympy.Symbol(var_name) in reverse_map:
                del reverse_map[sympy.Symbol(var_name)]

    # Operatoren für SymPy vorbereiten
    sympy_syntax = re.sub(r'\band\b', '&', masked_syntax, flags=re.IGNORECASE)
    sympy_syntax = re.sub(r'\bor\b', '|', sympy_syntax, flags=re.IGNORECASE)
    sympy_syntax = re.sub(r'\bnot\b', '~', sympy_syntax, flags=re.IGNORECASE)

    # --- 5. SymPy Logik-Kürzung & Re-Assembly ---
    final_syntax = ""
    try:
        expr = sympy.sympify(sympy_syntax, locals={k.name: k for k in reverse_map.keys()})
        simplified_expr = simplify_logic(expr, form='sop')
        
        if simplified_expr == true:
            final_syntax = "ALWAYS_TRUE"
        elif simplified_expr == false:
            final_syntax = "NEVER_POSSIBLE"
        else:
            final_syntax = str(simplified_expr)
            # Python Operatoren zurück zu SAP Operatoren
            final_syntax = final_syntax.replace('&', 'AND').replace('|', 'OR').replace('~', 'NOT ')
            # Entmaskieren
            for sym, original_cond in reverse_map.items():
                final_syntax = re.sub(fr'\b{sym.name}\b', original_cond, final_syntax)
                
    except Exception as e:
        # SMART FALLBACK: SymPy crasht (z.B. wegen fehlender Klammern). 
        # Wir geben trotzdem die SAP-Syntax zurück, in der wir True/False schon ersetzt haben!
        print(f"⚠️ AST Parser Fallback für '{cleaned_string[:30]}...': {e}")
        final_syntax = masked_syntax
        for sym, original_cond in reverse_map.items():
            final_syntax = re.sub(fr'\b{sym.name}\b', original_cond, final_syntax)

    # --- 6. DIE GHOST-BREADCRUMBS ANHÄNGEN ---
    if ghost_notes:
        notes_str = " | ".join(ghost_notes)
        final_syntax += f" /* GHOST-LOGIC: {notes_str} */"
        
    return final_syntax

# ==========================================
# 3. DER NEUE PRE-PROCESSOR
# ==========================================
def heavy_preprocessing(char_dict, known_canon):
    """
    1. Löscht Werte ohne Preconditions.
    2. Verbindet Arrays mit AND.
    3. Jagt die Syntax durch den AST Optimizer (inkl. Waschmaschine).
    """
    import copy
    char_copy = copy.deepcopy(char_dict)
    optimized_values = []
    
    for val in char_copy.get("values", []):
        deps = val.get("dependency", [])
        
        # SCHRITT 1: Löschen, wenn gar keine Abhängigkeit vorhanden ist
        if not deps:
            continue
            
        # SCHRITT 2: Alle Preconditions mit AND verknüpfen
        combined_syntax_parts = []
        for dep in deps:
            syntax = dep.get("syntax", "").strip()
            if syntax:
                combined_syntax_parts.append(f"({syntax})")
        
        if not combined_syntax_parts:
            continue
            
        combined_syntax = " AND ".join(combined_syntax_parts)
        
        # SCHRITT 3: AST Optimizer & Breadcrumbs
        optimized_syntax = optimize_sap_logic(combined_syntax, known_canon)
        
        val["dependency"] = [{"precondition": "OPTIMIZED_AST", "syntax": optimized_syntax}]
        optimized_values.append(val)
        
    char_copy["values"] = optimized_values
    return char_copy

# ==========================================
# 4. EXCEL DASHBOARD ERSTELLUNG
# ==========================================
def create_excel_dashboard(output_file, results_dict):
    """Erstellt das professionelle Entwickler-Dashboard in Excel."""
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    subheader_font = Font(name="Arial", size=12, bold=True, color="003366")
    text_font = Font(name="Arial", size=10)
    code_font = Font(name="Courier New", size=10, color="000000")
    code_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    table_header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for char_name, data in results_dict.items():
        ws = wb.create_sheet(title=char_name[:31]) 
        
        ws.merge_cells('A1:G1')
        ws['A1'] = f"MAS MIGRATIONS-VORSCHLAG: {char_name}"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 25

        ws['A3'] = "1. Herleitung (Gedankengang des Architekten)"
        ws['A3'].font = subheader_font
        ws['A4'] = data.get("herleitung", "Keine Herleitung verfügbar.")
        ws['A4'].font = text_font
        ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells('A4:G6')

        ws['A8'] = "2. Visuelle Variantentabelle (Für CU60)"
        ws['A8'].font = subheader_font
        
        start_row = 9
        headers = data.get("tabelle_kopf", [])
        for c_idx, header in enumerate(headers):
            cell = ws.cell(row=start_row, column=c_idx+1, value=str(header))
            cell.font = Font(bold=True)
            cell.fill = table_header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            
        rows = data.get("tabelle_zeilen", [])
        for r_idx, row in enumerate(rows):
            for c_idx, value in enumerate(row):
                cell = ws.cell(row=start_row + 1 + r_idx, column=c_idx+1, value=str(value))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

        code_start_row = start_row + len(rows) + 3
        ws.cell(row=code_start_row, column=1, value="3. Generierter AVC Constraint Code").font = subheader_font
        
        code_cell = ws.cell(row=code_start_row+1, column=1, value=str(data.get("avc_code", "// Kein Code generiert.")))
        code_cell.font = code_font
        code_cell.fill = code_fill
        code_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.merge_cells(start_row=code_start_row+1, start_column=1, end_row=code_start_row+15, end_column=7)

        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 20

    wb.save(output_file)
    print(f"📊 Excel Dashboard erfolgreich gespeichert unter: {output_file}")